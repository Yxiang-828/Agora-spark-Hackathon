from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
wb = Workbook()

HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
HEAD_FONT = Font(FONT, bold=True, color="FFFFFF", size=11)
ST_FILL = {"Built": "C6EFCE", "In progress": "FFEB9C", "Planned": "DDEBF7", "Later": "E7E6E6"}
ST_FONT = {"Built": "1E6B2F", "In progress": "7F6000", "Planned": "1F4E79", "Later": "595959"}
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

rows = [
("P0","Foundations","Freeze data contracts","Lock the shared shapes (action, directory, etc.) both halves build to","In progress","P1","-","All shared shapes agreed; changes need both owners"),
("P0","Foundations","4 authority tiers","Operator / Lead / Member / Guest over Mattermost roles","In progress","P1","-","A Member cannot approve; a Lead can"),
("P0","Foundations","Branding (Agora)","Plugin id + name renamed to Agora; swap logo/wordmark","In progress","P1","-","No 'Mattermost' wordmark shows post-login"),
("P1","The Room","Action cards","AI work shows as a rich card: main action + expandable sub-steps, live","Built","P1","contracts","Posting an action renders rich + collapsible, not raw text"),
("P1","The Room","Slash commands","/wrap /approve /pause /ai /claim","Built","P1","-","Each command fires the right action"),
("P1","Agents","Connector","Outbound connection, listens, replies in-thread, auto-reconnect","Built","P2","-","Human<->AI in a live thread, verified"),
("P1","Agents","AI adapters","Run Claude / Codex / Gemini headless with per-tool quirks","Built","P2","connector","All three answer in the room"),
("P1","Agents","Many bots per person","One owner can connect several AIs, each its own bot","Built","P2","adapters","One person, multiple online bots"),
("P1","Connecting","Pairing wizard","One-time code mints a bot + token","Built","P2","authority","Code to connected; states connected/failed/expired"),
("P1","Connecting","Downloadable bundle","Zip + per-OS launcher, no terminal typing","Built","P1","pairing","Download, double-click, connected"),
("P1","Connecting","Local supervisor","Runs all your agents, checks prereqs, restarts on crash","Built","P2","pairing","Agents survive a crash; clear skip reasons"),
("P1","Connecting","Connect/disconnect UI","Connect/disconnect any agent from the GUI + dashboard","Built","P1","liveness","GUI connect + disconnect both work"),
("P1","Engagement","Engagement rules","Replies on @mention / engaged thread; quiet in chatter","In progress","P2","connector","Follow-ups need no re-mention; no self-loops"),
("P1","Visibility","In-chat run log","Every command the AI runs shows as a step (no secrets)","In progress","P1","action cards","A multi-step task shows its steps in chat"),
("P2","The Room","Directory panel","Live list of who is connected (agent/owner/online)","In progress","P1","liveness","Survives restart; live list + status"),
("P2","Connecting","Liveness","Heartbeat to online/away/offline dot","Built","P2","connector","Disconnect flips to offline instantly"),
("P2","Engagement","Mute / channel off","Per-user mute; per-channel on/off; fail-closed","Built","P1","slash cmds","Muted = silent unless directly @mentioned"),
("P2","Authority","Authority gates","Who can approve/edit/mutate, enforced server-side","In progress","P1","auth tiers","Gates fail-closed; verified per tier"),
("P2","The Brain","The Gate to Archive","/wrap to proposal to authorized human approves to saved","Built","P1","auth gates","Approved case is saved; provenance server-derived"),
("P2","The Brain","The Dictionary","Searchable problem to cause to fix to provenance index","Built","P1","Gate","An approved case is a searchable entry"),
("P2","The Brain","Tiered memory","Conversation / facts / skills / findings kept in tiers","In progress","P1","auth tiers","Gate is the only durable writer"),
("P2","Skills","Layered skills","Core/workplace/personal/grown; room serves a skill list on connect","In progress","P2","connector","Agent gets the workplace skill list on connect"),
("P2","Skills","Skills gate","Room re-checks every skill server-side, rejects with reasons","Built","P1","skills","Bad skills rejected with a clear reason"),
("P2","Channels","Default channels","Setup creates welcome/features/code-review + welcome post","Built","P1","slash cmds","/agora setup builds the channels"),
("P2","Channels","Per-channel AI toggle","Turn the AI on/off in a channel","Built","P1","slash cmds","Off = bot ignores that channel"),
("P2","Quality","Cross-OS check harness","Auto-verify a skill works (or fails gracefully) on Win/macOS/Linux","In progress","P2","skills gate","Same skill passes on all 3, recorded"),
("P3","Orchestration","Case & task tracking","One coordinator owns start/stop/resume of AI tasks","In progress","P1","contracts","Tasks resume after interrupt; checkpointed"),
("P3","Orchestration","Resource locks","Two tasks cannot mutate the same thing; read-only runs free","Planned","P1","case/task","Conflicting mutations serialize, no auto-retry"),
("P3","Agents","Agent roles","An agent holds a job: orchestrator/debug/audit/CI/general","Planned","P2","auth tiers","A bot can be assigned a role + scoped skills"),
("P3","Agents","Orchestrator agent","Splits a request, dispatches to agents, merges results","Planned","P2","roles + case/task","One request fans out and comes back merged"),
("P3","Agents","Private workspace","Each agent works in its own folder/config, separate from shared code","Built","P2","adapters","Agent picks up its own folder config"),
("P3","Codespace IDE","Host-backed git","Edit a real repo on a member machine via the room","Built","P3","connector + auth","Browse/edit/commit/push a real repo"),
("P3","Codespace IDE","Real-time editing","Several people edit the SAME file at once, merges live","Planned","P3","host-backed git","Two browsers edit one file, it converges"),
("P3","Codespace IDE","Live cursors","See where teammates are typing","Planned","P3","real-time editing","Cursors/selection show per user"),
("P3","Codespace IDE","Save to git","Live edits flush to disk; commit/push on demand","Planned","P3","real-time editing","Edits land on the host git repo"),
("P3","Codespace IDE","Live edit attribution (blame)","Every line shows who last changed it - a human user or which AI bot - from the realtime edit origin (not just at commit)","Planned","P3","real-time editing","Hovering a line shows who last edited it + when"),
("P3","Codespace IDE","Author coloring overlay","Toggle to tint text by author, like Google Docs 'edits by person'","Planned","P3","live edit attribution","Toggle colors each author text distinctly"),
("P3","Codespace IDE","Presence avatars + follow","See who is in the file; click an avatar to follow their viewport/cursor","Planned","P3","live cursors","Avatars show; follow jumps to their cursor"),
("P3","Visibility","Error transparency","On failure, chat shows which command failed + the reason","Planned","P1","run log","No silent failures; typed reason in chat"),
("P3","Visibility","Live state","Card shows running/done/failed/paused + health dot","In progress","P1","run log","State always visible to the user"),
("P3","The Room","Console panels","RHS: live action / brain browser / approval queue (polished)","In progress","P1","action cards + brain","A promotion can be approved from the panel"),
("P4","Skills","Cross-OS skills","One skill resolves the right tool per OS","In progress","P2","skills","ssh-access runs from Windows AND Linux"),
("P4","Skills","ssh-access + fleet-ops","The skills for the read-only robot-ops demo","Planned","P2","cross-OS skills","Read-only ops diagnosis runs via the skill"),
("P4","Agents","Dedicated channel agents","Debug / audit (cited) / CI-CD agents, each in its channel","Planned","P2","roles","Each engages on its channel, scoped skills"),
("P4","Orchestration","Agent overlap handling","Orchestrator spots two agents on the same area, serializes/warns","Planned","P1","locks + orchestrator","Overlap gets serialized or a public warning"),
("P4","Codespace IDE","Directory tree","Folder sidebar: expand, create, rename, delete files/folders","In progress","P3","host-backed git","Full file/folder CRUD over host disk"),
("P4","Codespace IDE","Rules engine","Enforce edit/commit/push rights, protected paths, commit msgs","Planned","P3","real-time editing + auth","A rule reject gives a clear inline reason"),
("P4","Codespace IDE","Speed & comfort","Instant local typing; no lost keystrokes on reconnect; offline banner","Planned","P3","real-time + cursors","Under 50ms local echo; clean reconnect merge"),
("P4","Codespace IDE","Agent writes here","A channel AI writes code into the shared codespace, live","Built","P2","host-backed git + bind","Agent code blocks appear in the editor"),
("P4","Codespace IDE","Who-wrote-this + why","Select a range to see last editor, time, and (if AI) the case/prompt that produced it - links back to chat","Planned","P3","live edit attribution + run log","An AI-written line links back to its chat case"),
("P4","Codespace IDE","Version history (timeline)","Google-Docs-style scrubber of named snapshots; open the file/codespace at any past point","Planned","P3","real-time editing","You can open and read any past version"),
("P4","Codespace IDE","Diff between versions","See what changed, and by whom, between two points in history","Planned","P3","version history","Inline/side-by-side diff with per-author marks"),
("P4","Codespace IDE","Restore / revert","Restore a past version, or revert one author change; rules-gated","Planned","P3","version history + rules engine","Restore is recorded as a new attributed change"),
("P4","Codespace IDE","Activity feed","Chronological per-file/codespace log: edits, commits, create/rename/delete, by whom","Planned","P3","live edit attribution","A readable 'who did what when' timeline"),
("P4","Codespace IDE","Commit co-authorship","On commit, record the humans + AIs whose edits are in it (Co-authored-by/manifest) so downstream git blame keeps the attribution","Planned","P3","save to git + live edit attribution","A commit lists its real human + AI co-authors"),
("P4","Codespace IDE","Durable attribution log","Append-only server-side record of every edit/commit with author + time (feeds the audit agent)","Planned","P3","live edit attribution","Every edit/commit is in an immutable server log"),
("P4","Codespace IDE","Inline comments","Google-Docs-style comments on a line/range, threaded and resolvable","Planned","P3","real-time editing","Comment on a selection and resolve it"),
("P4","Codespace IDE","@mention in code/comments","Ping a teammate or an AI from a comment; it posts to chat with a deep-link","Planned","P3","inline comments","@mention notifies + links to the exact spot"),
("P4","Codespace IDE","Soft file/range locks","Claim a file or region to avoid churn; others get a non-blocking warning","Planned","P3","rules engine","A locked region warns others before they edit"),
("P4","Codespace IDE","Find across codespace","Search all files in the codespace (server-side over host disk)","Planned","P3","host-backed git","A query returns hits across files with paths"),
("P4","Codespace IDE","Review before commit","Stage edits, review the diff, approve, then commit (rules + authority gated)","Planned","P3","diff between versions + rules engine","A commit can require a reviewer"),
("P4","Connecting","Cross-agent delegate","One person AI hands off to another (loop-protected)","Planned","P2","pairing + orchestrator","Two agents delegate with hop-limit + cooldown"),
("P4","The Brain","Distiller","Turns a case into a proposal; suggests update vs new","Planned","P2","tiered memory","Proposes UPDATE vs CREATE, no duplicates"),
("P4","Quality","End-to-end probes","Pass/fail test of the real flow per feature","In progress","P1","-","Each feature has a live probe"),
("P4","Quality","Clear error contracts","Skills/code return proper typed errors on bad input","In progress","P2","probes","Bad input to enumerated typed error, no crash"),
("P4","Quality","Graceful-failure tests","Verify reconnects, safe defaults, clean partial recovery","In progress","P1","probes","Reconnect/backoff + fail-closed verified"),
("P3","Comms","Voice channels","On-demand or persistent voice rooms per channel (WebRTC), built on Mattermost Calls","Planned","P1","-","Two members talk by voice in a channel"),
("P4","Comms","Screen share (low latency)","WebRTC screen share in a call, tuned for sub-second latency","Planned","P1","voice channels","Viewers see a shared screen with sub-second lag"),
("P4","Comms","Demo mode","Present screen + voice to the channel; one-click record","Planned","P1","screen share","Run a live demo others can watch"),
("P4","Comms","Call recording + transcript","Record a call/demo, transcribe it, attach to the case/brain","Later","P2","demo mode","A demo is saved with a searchable transcript"),
("P5","Comms","AI in the call","An AI joins a call: watches the shared screen (vision) + answers by voice","Later","P2","screen share","The AI answers about what is on the shared screen"),
("P4","Frontend Control","Agent browser session","A managed, sandboxed real browser (Chromium/CDP) the agent drives per task, lifecycle owned by the coordinator - not ad-hoc Playwright spawns","Planned","P2","case/task","Agent opens a page in a managed cross-OS browser"),
("P4","Frontend Control","DOM + accessibility read","Agent reads a structured semantic snapshot (DOM + ARIA roles), token-efficient - not raw HTML or screenshots","Planned","P2","agent browser session","Agent gets the page as a structured tree"),
("P4","Frontend Control","Visual + element map","Screenshot plus an element-to-bounding-box map so the agent reasons visually AND structurally together","Planned","P2","DOM + accessibility read","Agent links a screenshot region to a DOM node"),
("P4","Frontend Control","Semantic interaction intents","High-level actions (click 'Submit', fill 'Email') resolved by role/label/text - self-healing, not brittle CSS/XPath (beats Playwright)","Planned","P2","DOM + accessibility read","Actions work without hardcoded selectors"),
("P4","Frontend Control","Live console/network/errors","Agent sees JS console, network calls, and runtime errors as structured events","Planned","P2","agent browser session","Agent reads console + failed requests live"),
("P4","Frontend Control","Component/state inspection","Framework-aware read of the React/Vue component tree + props/state (beats Codex GUI debugging)","Planned","P2","DOM + accessibility read","Agent reads a live component prop/state value"),
("P4","Frontend Control","Edit-preview-verify loop","Agent edits frontend code in the codespace, the preview hot-reloads, the agent re-reads the DOM to confirm the change","Planned","P2","semantic interaction intents + host-backed git","Agent confirms a UI change without manual Playwright"),
("P4","Frontend Control","Agent control panel","A room panel where a human watches + steers the agent in the browser: same DOM/console view, one-click hand-off","Planned","P1","DOM + accessibility read","A human sees and can take over the agent browser"),
("P4","Frontend Control","Element picker (human to agent)","Human clicks an element in the live view to point the agent at it ('fix this button')","Planned","P1","agent control panel","Clicking an element hands its node to the agent"),
("P4","Frontend Control","Live preview / demo host","A shareable live URL of the frontend the agent is building","Planned","P3","host-backed git","A preview URL renders the current frontend"),
("P5","Frontend Control","Record interaction as a test","The agent actions become a durable, attributed, replayable test (beats hand-written Playwright)","Later","P2","semantic interaction intents","A run is saved as a re-runnable test"),
("P5","Later","Cross-use grants","Time-boxed consent to use another AI (Ask/Delegate/Drive)","Later","P1","auth gates + delegate","Grant token authorizes; creds never transfer"),
("P5","Later","Skill graduation","Lead promotes a grown skill to the shared set","Later","P2","skills + compat","Graduates only after passing on all 3 OSes"),
("P5","Later","Semantic search","Find related past cases by meaning (vector)","Later","P1","tiered memory","Recall related cases without exact keywords"),
("P5","Later","Sessions & modes","Group an AI session w/ mode (solo/co-op/debate/drive)","Later","P1","case/task","A session holds participants + a mode"),
("P5","Later","Timeline & decisions","Per-thread log of what happened + why","Later","P1","run log + console","Replayable per-thread decision log"),
("P5","Later","Codespace terminal","A safe sandboxed terminal per codespace","Later","P3","host-backed git","Run commands in a per-codespace sandbox"),
("P5","Later","Code language features (LSP)","Go-to-definition, hover types, lint in the editor via a language server","Later","P3","real-time editing","Go-to-def + hover work in a file"),
("P5","Later","CI/CD check agent","Watches a repo, runs build+test+lint+compat, posts pass/fail","Later","P2","channel agents + probes","Posts a pass/fail action with logs in chat"),
("P5","Later","Full white-label","Login page, favicon, every string","Later","P1","branding","Zero Mattermost strings anywhere"),
("P5","Later","Scheduler","Time/condition triggers, restart-safe, missed-fire policy","Later","P2","-","at/every/after/deadline fire reliably"),
]

ws = wb.active
ws.title = "Feature Checklist"
headers = ["Done?", "Phase", "Area", "Feature", "What it does", "Status", "Suggested person", "Depends on", "Done when"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(1, c)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

for r in rows:
    phase, area, feat, what, status, person, dep, done = r
    tick = "Y" if status == "Built" else ""
    ws.append([tick, phase, area, feat, what, status, person, dep, done])

last = ws.max_row
for r in range(2, last + 1):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(r, c)
        cell.border = BORDER
        cell.font = Font(FONT, size=10)
        cell.alignment = WRAP if c in (3, 4, 5, 8, 9) else CTR
    d = ws.cell(r, 1)
    if d.value == "Y":
        d.fill = DONE_FILL
        d.font = Font(FONT, bold=True, color="1E6B2F", size=11)
    st = ws.cell(r, 6).value
    if st in ST_FILL:
        ws.cell(r, 6).fill = PatternFill("solid", fgColor=ST_FILL[st])
        ws.cell(r, 6).font = Font(FONT, bold=True, color=ST_FONT[st], size=10)

W = {"A": 7, "B": 7, "C": 15, "D": 22, "E": 46, "F": 13, "G": 10, "H": 22, "I": 40}
for k, v in W.items():
    ws.column_dimensions[k].width = v
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:I%d" % last
ws.row_dimensions[1].height = 30

def adddv(col, opts):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(opts), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("%s2:%s%d" % (col, col, last))

adddv("A", ["Y", ""])
adddv("F", ["Built", "In progress", "Planned", "Later"])
adddv("G", ["P1", "P2", "P3"])
adddv("B", ["P0", "P1", "P2", "P3", "P4", "P5"])

g = wb.create_sheet("Guide")
guide = [
("Agora - Feature Checklist", "title"),
("How to read this", "h"),
("Done?", "Mark Y when the feature is finished. Already-built features are pre-marked."),
("Phase", "Build order. Each phase builds on the one before; within a phase, work runs in parallel."),
("Status", "Built = working & verified. In progress = partly there. Planned = v1.0. Later = after v1.0."),
("Suggested person", "Who could build it (a suggestion, not a fixed assignment) - see below."),
("Depends on", "What must exist first."),
("Done when", "The acceptance check - when it counts as finished."),
("", "sp"),
("Suggested people (3)", "h"),
("P1", "The Room - the Mattermost plugin: chat surface, Console, authority, the brain/Dictionary, orchestration."),
("P2", "The Agents - everything on the user machine + the learning loop: connector, agent roles, skills, growth."),
("P3", "The Codespace IDE - the shared real-time editor: live code sync, git, directory tree, rules."),
("", "sp"),
("The phases", "h"),
("P0", "Foundations - freeze the contracts, authority tiers, branding."),
("P1", "Spine - room and agents talk: action cards, connector, pairing, engagement, run log."),
("P2", "Trust & knowledge - authority gates, directory, the Gate to Dictionary, skills, cross-OS checks."),
("P3", "Orchestration & IDE core - case/task tracking, agent roles, orchestrator, real-time codespace, error transparency."),
("P4", "The demo & headline - ssh-access, debug/audit/CI agents, overlap handling, IDE tree+rules+speed, probes."),
("P5", "Later - debate mode, drive grants, semantic search, scheduler, terminal, full white-label."),
("", "sp"),
("v1.0 is done when", "h"),
("", "A member local AI joins from any OS, then @their-agent runs a read-only ops check via ssh-access, it shows as a live action card, then /wrap, a Lead approves, and it lands in the shared Dictionary. Plus: two people editing one file live in the shared codespace, merged and saved to git."),
("", "sp"),
("The rules every feature follows", "h"),
("1", "Works on every OS, or fails with a clear typed error - never a crash, never a silent single-OS assumption."),
("2", "Secrets never reach the AI - resolved only at use time, never in prompts/messages/logs."),
("3", "Nothing is saved to the shared brain without a human approving it."),
("4", "Risky actions are guarded - read-only is free; changes lock, are not blindly retried; high-impact needs confirm + auto-undo."),
("5", "No silent failures - every error is surfaced with a clear, specific reason."),
("6", "No made-up critique - claims backed by the code or a cited practice."),
("7", "Everything is tested - unit tests AND a real end-to-end check. Done = the demo works."),
]
g.column_dimensions["A"].width = 20
g.column_dimensions["B"].width = 95
gr = 1
for item in guide:
    if item[1] == "title":
        g.cell(gr, 1, item[0]).font = Font(FONT, bold=True, size=16, color="1F3A5F")
        gr += 2
        continue
    if item[1] == "h":
        g.cell(gr, 1, item[0]).font = Font(FONT, bold=True, size=12, color="FFFFFF")
        g.cell(gr, 1).fill = HEAD_FILL
        g.cell(gr, 2).fill = HEAD_FILL
        gr += 1
        continue
    if item[1] == "sp":
        gr += 1
        continue
    g.cell(gr, 1, item[0]).font = Font(FONT, bold=True, size=10)
    g.cell(gr, 1).alignment = Alignment(vertical="top")
    g.cell(gr, 2, item[1]).font = Font(FONT, size=10)
    g.cell(gr, 2).alignment = WRAP
    gr += 1

import itertools
saved = None
for name in ["Agora-Features-Checklist.xlsx"] + ["Agora-Features-Checklist-v%d.xlsx" % i for i in itertools.count(2)]:
    try:
        wb.save(name)
        saved = name
        break
    except PermissionError:
        if name.endswith("v9.xlsx"):
            raise
        continue
print("saved rows:", len(rows), "->", saved, "(LOCKED ones skipped)" if saved != "Agora-Features-Checklist.xlsx" else "")
